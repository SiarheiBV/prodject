from datetime import datetime
from openpyxl import Workbook, load_workbook
import os


def create_category(category: str, parameters: list, connector):
    try:
        data = connector.read_file()
    except FileNotFoundError:
        data = {}

    if category not in data:
        data[category] = parameters
        connector.write_file(data)
        print("Category added")
    else:
        raise NameError("Category already exists")


def add_product(category: str, quantity: int, connector_db, cat_json):
    cat_list = cat_json.read_file()
    try:
        data = connector_db.read_file()
    except FileNotFoundError:
        data = {}

    if category not in cat_list:
        raise NameError("no such category")

    parameters = {}
    for parameter in cat_list[category]:
        value = input(f"Enter value for '{parameter}': ")
        parameters[parameter] = value

    existing_product = None
    for product in data.get(category, []):
        if product.get("Id") == parameters.get("Id"):
            existing_product = product
            break

    if existing_product:
        existing_product["quantity"] += quantity
        existing_product["updated_at"] = str(datetime.now().strftime("%d.%m.%Y"))
        connector_db.write_file(data)
        print("Product updated.")
    else:
        parameters["category"] = category
        parameters["quantity"] = quantity
        parameters["created_at"] = str(datetime.now().strftime("%d.%m.%Y"))
        parameters["updated_at"] = parameters["created_at"]
        data.setdefault("GOODS", []).append(parameters)
        connector_db.write_file(data)
        print("New product added.")


def get_all(category, min_date=None, max_date=None, connector_db=None):
    data = connector_db.read_file()
    result = []

    for _, items in data.items():
        for item in items:
            if not category or category == item["category"]:
                create_date_obj = datetime.strptime(item["updated_at"], "%d.%m.%Y")
                if (not min_date or create_date_obj >= datetime.strptime(min_date, "%d.%m.%Y")) and (not max_date or create_date_obj <= datetime.strptime(max_date, "%d.%m.%Y")):
                    result.append({"category": item["category"], "id": item["Id"], "model": item["Model"], "price": item["Price"]})
    if result:
        print(*result)
    else:
        print("No matching products found.")


def get_product(pr_code, connector_db):
    data = connector_db.read_file()

    flag = True
    for _, items in data.items():
        for item in items:
            if item["Id"] == pr_code:
                flag = True
                if int(item["quantity"]) > 0:
                    print(f"You can buy {item['Model']}. Quantity: {item['quantity']}")
                    break
                else:
                    print("Product is out of stock")
                    break
            else:
                flag = False
    if not flag:
        print("product code is not defined")


def place_an_order(basket, connector_db, orders_file):
    data = connector_db.read_file()

    shopping_list = []
    total_cost = 0

    for code, quantity in basket.items():
        flag = False
        for item in data["GOODS"]:
            if item['Id'] == code:
                flag = True
                if int(item['quantity']) >= int(quantity):
                    item["quantity"] = int(item["quantity"]) - int(quantity)
                    current_cost = int(item["Price"]) * int(quantity)
                    total_cost += current_cost
                    shopping_list.append({"id": item['Id'], "category": item['category'], "name": item['Model'], "current_cost": current_cost, "quantity": quantity})
                else:
                    print(f"Not enough products with code {code}")
        if not flag:
            print(f"Product with code {code} is not defined in the store")
    connector_db.write_file(data)
    print("Your purchases:", *shopping_list, "total cost:", total_cost)

    try:
        orders = orders_file.read_file()
        count = len(orders["ORDERS"])
        print(count)
    except FileNotFoundError:
        orders = {}
        count = 0

    order = {
        "id": count + 1,
        "items": shopping_list,
        "total_cost": total_cost,
        "created_at": str(datetime.now().strftime("%d.%m.%Y"))
    }
    orders.setdefault("ORDERS", []).append(order)
    orders_file.write_file(orders)


def get_status(min_date=None, max_date=None, connector_db=None):
    data = connector_db.read_file()
    result = []

    for items in data.values():
        for item in items:
            create_date_obj = datetime.strptime(item["created_at"], "%d.%m.%Y")
            if (not min_date or create_date_obj >= datetime.strptime(min_date, "%d.%m.%Y")) and (not max_date or create_date_obj <= datetime.strptime(max_date, "%d.%m.%Y")):
                result.append(item)

    file_path = 'data/status.xlsx'

    if not os.path.isfile(file_path):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Categories'
        workbook.create_sheet('Products')
        workbook.create_sheet('Orders')
        workbook.create_sheet('Metrics')
    else:
        workbook = load_workbook(filename=file_path)

    worksheet = workbook['Categories']
    worksheet['A1'] = 'Category'
    worksheet['B1'] = 'Quantity'
    worksheet['C1'] = 'Revenue'

    categories = {}
    for items in result:
        for item in items["items"]:
            category = item['category']
            if category not in categories:
                categories[category] = {'quantity': 0, 'revenue': 0}
            categories[category]['quantity'] += int(item['quantity'])
            categories[category]['revenue'] += int(item['current_cost']) * int(item['quantity'])

    row = 2
    for category, stats in categories.items():
        worksheet.cell(row=row, column=1, value=category)
        worksheet.cell(row=row, column=2, value=stats['quantity'])
        worksheet.cell(row=row, column=3, value=stats['revenue'])
        row += 1

    worksheet = workbook['Products']
    worksheet['A1'] = 'Id'
    worksheet['B1'] = 'Model'
    worksheet['C1'] = 'Quantity'

    product_quantities = {}
    for items in result:
        for item in items["items"]:
            product_code = item['id']
            if product_code not in product_quantities:
                product_quantities[product_code] = {'name': item["name"], "quantity": 0}
            product_quantities[product_code]["quantity"] += int(item['quantity'])

    row = 2
    for code, stats in product_quantities.items():
        worksheet.cell(row=row, column=1, value=code)
        worksheet.cell(row=row, column=2, value=stats['name'])
        worksheet.cell(row=row, column=3, value=stats["quantity"])
        row += 1

    worksheet = workbook['Orders']
    worksheet['A1'] = 'Order Id'
    worksheet['B1'] = 'Name Product'
    worksheet['C1'] = 'Total'
    worksheet['D1'] = 'Order date'

    sort_orders = sorted(result, key=lambda x: int(x['total_cost']), reverse=True)
    row = 2
    for items in sort_orders:
        for item in items['items']:

            worksheet.cell(row=row, column=1, value=items['id'])
            worksheet.cell(row=row, column=2, value=item['name'])
            worksheet.cell(row=row, column=3, value=items['total_cost'])
            worksheet.cell(row=row, column=4, value=items['created_at'])
        row += 1

    worksheet = workbook['Metrics']
    worksheet['A1'] = 'Total Cost'
    worksheet['B1'] = 'Total Goods'
    worksheet['C1'] = 'Popular Category'
    worksheet['D1'] = 'Popular Product'

    total_cost = sum(int(item['total_cost']) for item in result)
    total_goods = sum(int(item['quantity']) for items in result for item in items["items"])
    categories = {}
    for items in result:
        for item in items['items']:
            category = item['category']
            quantity = item['quantity']
            if category not in categories:
                categories[category] = int(quantity)
            categories[category] += int(quantity)

    popular_category = max(categories, key=categories.get)

    products = {}
    for order in result:
        for item in order['items']:
            product = item['name']
            quantity = item['quantity']
            if product not in products:
                products[product] = int(quantity)
            products[product] += int(quantity)
    popular_product = max(products, key=products.get)

    metrics = {"total_cost": total_cost, "total_goods": total_goods, "popular_category": popular_category, "popular_product": popular_product}

    row = 2
    for item in metrics.items():
        worksheet.cell(row=row, column=1, value=metrics['total_cost'])
        worksheet.cell(row=row, column=2, value=metrics['total_goods'])
        worksheet.cell(row=row, column=3, value=metrics['popular_category'])
        worksheet.cell(row=row, column=4, value=metrics['popular_product'])

    workbook.save(file_path)
    print(f"data successfully saved to file {file_path}")
    # try:
    #     workbook.save(file_path)
    #     print(f"Данные успешно сохранены в файл {file_path}")
    # except Exception as e:
    #     print(f"Ошибка сохранения файла")


"""г) На четвёртом листе должны выводиться метрики: общая выручка за выбранный период, общее количество товаров, самая популярная категория,
самый популярный товар."""
